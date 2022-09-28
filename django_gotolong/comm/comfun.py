import openpyxl
import pandas as pd
import re

import io
import random

import os
import PyPDF2

from pikepdf import Pdf


def comm_func_ticker_match(ticker, amfi_rank_dict, dematsum_list):
    if (ticker in amfi_rank_dict and amfi_rank_dict[ticker] <= 500) \
            or ticker in dematsum_list:
        return True
    else:
        return False


def comm_mf_subcat(fund_name, benchmark_name):
    if re.search(r'Flexi', fund_name, re.IGNORECASE):
        fund_subcat = 'F-500'
    elif re.search(r'Large', fund_name, re.IGNORECASE) \
            or re.search(r'Blue.*chip', fund_name, re.IGNORECASE) \
            or re.search(r'Frontline', fund_name, re.IGNORECASE):
        if re.search(r'Mid', fund_name, re.IGNORECASE):
            fund_subcat = 'LM-250'
        else:
            fund_subcat = 'L-100'
    elif re.search(r'Mid', fund_name, re.IGNORECASE):
        fund_subcat = 'M-150'
    elif re.search(r'Small', fund_name, re.IGNORECASE):
        fund_subcat = 'S-250'
    elif re.search(r'Multi', fund_name, re.IGNORECASE):
        fund_subcat = 'U-Multi'
    elif re.search(r'Value', fund_name, re.IGNORECASE):
        fund_subcat = 'U-Value'
    elif re.search(r'Dividend', fund_name, re.IGNORECASE):
        fund_subcat = 'U-Dividend'
    elif re.search(r'Index', fund_name, re.IGNORECASE):
        fund_subcat = 'U-Index'
    elif re.search(r'ETF', fund_name, re.IGNORECASE):
        fund_subcat = 'U-ETF'
    else:
        fund_subcat = 'U-Unknown'
        """
        if re.search('500', benchmark_name):
            fund_subcat = 'F-500'
        elif re.search('100', benchmark_name):
            fund_subcat = 'L-100'
        elif re.search('150', benchmark_name):
            fund_subcat = 'M-150'
        elif re.search('250', benchmark_name):
            fund_subcat = 'S-250'
        else:
            pass
        """
    return fund_subcat


# one parameter named request
def comm_func_upload(request, template, columns_list, list_url_name, ignore_top_lines=0):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    # template = "imf/BrokerIcidirMf_list.html"
    # data = BrokerIcidirMf.objects.all()

    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template)

    req_file = request.FILES['file']

    # let's check if it is a csv file

    if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
        # get worksheet name
        # print('temporary file path:', req_file.temporary_file_path)
        print(req_file)

        if True:
            wb = openpyxl.load_workbook(req_file)
            print(wb.sheetnames)
            sheet_name = wb.sheetnames[0]
            print(sheet_name)
            ws = wb[sheet_name]
            df = pd.DataFrame(ws.values)
        else:
            xl = pd.ExcelFile(req_file)
            if debug_level > 0:
                print(xl.sheet_names)
            # single worksheet - Data
            sheet_name = xl.sheet_names[0]
            df = xl.parse(sheet_name)

        # can be 'Data'
        # can be 'Average MCap Jan Jun 2020'
        # if sheet_name != 'fund-performance':
        #    print("sheet name changed to", sheet_name)

        # ignore top 6 line : Value Research, Fund Performance
        # remove top six line from dataframe

        # ignore the top 1 line
        # df = df.iloc[1:]
        # ignore top number of lines
        if ignore_top_lines != 0:
            df = df.iloc[ignore_top_lines:]

        if debug_level > 0:
            print("old columns : ")
            print(df.columns)

        # TBD
        # change column name of data frame
        # df.columns = columns_list

        if debug_level > 0:
            print("new columns : ")
            print(df.columns)

        # Keep only top 1000 entries
        # df = df.iloc[:1000]

        # round avg_mcap
        # df = df.round({'avg_mcap' : 1})
        # covert to numeric
        # df[["avg_mcap"]] = df[["avg_mcap"]].apply(pd.to_numeric)
        # df[["daily_aum"]] = df[["daily_aum"]].astype(int)

        # drop columns that are not required
        # skip_columns_list = ["none"]
        # df.drop(skip_columns_list, axis=1, inplace=True)

        data_set = df.to_csv(header=True, index=False)

    elif req_file.name.endswith('.pdf') or req_file.name.endswith('.PDF'):

        print(req_file)

        if 'password' in request.POST:
            pan_card_number = request.POST.get('password')
        else:
            pan_card_number = ''

        pdf_file_obj = ''
        if pan_card_number != '':
            pdf_reader_name = 'pikepdf'
            # decrypt using password
            new_pdf = Pdf.new()
            # NOTE: How will it work in production?
            # Can we make it work without saving the file.
            # It works only if the file has been removed by us

            decrypted_file_path = 'udepcas_output_' + pan_card_number + '.pdf'
            with Pdf.open(req_file, password=pan_card_number) as pdf:
                # remove before saving
                if os.path.exists(decrypted_file_path):
                    os.remove(decrypted_file_path)
                pdf.save(decrypted_file_path)
                pdf_file_obj = open(decrypted_file_path, 'rb')
                # pdf.pages[0].Contents.page_contents_coalesce()
                # pdf_file_obj = io.BytesIO(pdf.pages[0].Contents.get_stream_buffer())
        else:
            pdf_reader_name = 'PyPDF2'
            # create file object variable
            # opening method will be rb

            pdf_file_obj = req_file.open()

        print('used pdf reader name', pdf_reader_name)

        # create reader variable that will read the pdffileobj
        pdf_reader = PyPDF2.PdfFileReader(pdf_file_obj)

        # This will store the number of pages of this pdf file
        x = pdf_reader.numPages
        print('num pages', x)

        data_set = []
        for page_num in range(x):
            if debug_level > 1:
                print('------page ----begin---', page_num)
            # create a variable that will select the selected number of pages
            page_obj = pdf_reader.getPage(page_num)

            # (x+1) because python indentation starts with 0.
            # create text variable which will store all text datafrom pdf file
            # text = pageobj.extractText()
            text = page_obj.extract_text()

            for line in text.split('\n'):
                data_set.append(line)

            if debug_level > 1:
                print('------page ----end---', page_num)

        # must close the file before renaming
        pdf_file_obj.close()

        if pdf_reader_name == "pikepdf":

            # remove before leaving
            if os.path.exists(decrypted_file_path):
                print('Trying to remove file', decrypted_file_path)
                try:
                    os.remove(decrypted_file_path)
                    print('Removed file', decrypted_file_path)
                except OSError:
                    print('Failed to remove file', decrypted_file_path)

            else:
                print(decrypted_file_path + ' not exist')
        else:
            print('used different pdf_reader_name', pdf_reader_name)

    if req_file.name.endswith('.csv'):
        data_set = req_file.read().decode('UTF-8')

    if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') \
            or req_file.name.endswith('.xlsx') or req_file.name.endswith('.pdf') or req_file.name.endswith('.PDF')):
        messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV/PDF FILE.')
        return HttpResponseRedirect(reverse(list_url_name))

    return data_set
