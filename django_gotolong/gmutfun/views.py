# Create your views here.

from .models import Gmutfun

from django.views.generic.list import ListView

from django.db.models import (Count, Q)

from django.urls import reverse
from django.http import HttpResponseRedirect
import urllib3
import math

import csv
import io
import re

import openpyxl

import pandas as pd
import numpy as np

from django_gotolong.comm import comfun
from django_gotolong.lastrefd.models import Lastrefd, lastrefd_update

from itertools import zip_longest


class GmutfunListView(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_AUM_CapBox(ListView):
    model = Gmutfun

    def get_queryset(self):

        self.queryset = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Active')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum', 'gmutfun_benchmark'). \
            filter(gmutfun_aum__gte=100)

        return self.queryset

    # @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(GmutfunListView_Active_AUM_CapBox, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url

        min_score_grade = 0

        qs_flexi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_scheme__contains='Flexi')). \
            filter(Q(gmutfun_benchmark__contains='500 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        qs_large = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='100 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        qs_lami = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='LM-250')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        qs_mid = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Midcap 150 Total Return Index') |
                   Q(gmutfun_benchmark__contains='150 MidCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        qs_small = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Smallcap 250 Total Return Index') |
                   Q(gmutfun_benchmark__contains='250 SmallCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        qs_multi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='U-Multi')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        qs_value = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='U-Value')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        qs_divi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='U-Dividend')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_aum', '-gmutfun_alpha_pct')

        flex_list = []
        large_list = []
        lami_list = []
        mid_list = []
        small_list = []
        multi_list = []
        value_list = []
        divi_list = []

        # select at max 10
        max_count = 10

        cur_idx = 0
        for q1 in qs_flexi:
            cur_idx += 1
            if cur_idx > max_count:
                break
            flex_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_large:
            cur_idx += 1
            if cur_idx > max_count:
                break
            large_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_lami:
            cur_idx += 1
            if cur_idx > max_count:
                break
            lami_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_mid:
            cur_idx += 1
            if cur_idx > max_count:
                break
            mid_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_small:
            cur_idx += 1
            if cur_idx > max_count:
                break
            small_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_multi:
            cur_idx += 1
            if cur_idx > max_count:
                break
            multi_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_value:
            cur_idx += 1
            if cur_idx > max_count:
                break
            value_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_divi:
            cur_idx += 1
            if cur_idx > max_count:
                break
            divi_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        context["all_list"] = list(
            zip_longest(flex_list, large_list, lami_list, mid_list, small_list,
                        multi_list, value_list, divi_list, fillvalue='-'))

        return context

    def get_template_names(self):
        app_label = 'gmutfun'
        template_name_first = app_label + '/' + 'gmutfun_capbox_active_list.html'
        template_names_list = [template_name_first]
        return template_names_list


class GmutfunListView_Active_Select_CapBox(ListView):
    model = Gmutfun

    def get_queryset(self):

        self.queryset = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Active')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        return self.queryset

    # @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(GmutfunListView_Active_Select_CapBox, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url

        min_score_grade = self.kwargs['score_grade']

        qs_flexi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_scheme__contains='Flexi')). \
            filter(Q(gmutfun_benchmark__contains='500 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_large = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='100 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_lami = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='LM-250')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_mid = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Midcap 150 Total Return Index') |
                   Q(gmutfun_benchmark__contains='150 MidCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_small = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Smallcap 250 Total Return Index') |
                   Q(gmutfun_benchmark__contains='250 SmallCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_multi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='U-Multi')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_value = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='U-Value')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_divi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_subtype='U-Dividend')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=min_score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        flex_list = []
        large_list = []
        lami_list = []
        mid_list = []
        small_list = []
        multi_list = []
        value_list = []
        divi_list = []

        # select at max 10
        max_count = 10

        cur_idx = 0
        for q1 in qs_flexi:
            cur_idx += 1
            if cur_idx > max_count:
                break
            flex_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_large:
            cur_idx += 1
            if cur_idx > max_count:
                break
            large_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_lami:
            cur_idx += 1
            if cur_idx > max_count:
                break
            lami_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_mid:
            cur_idx += 1
            if cur_idx > max_count:
                break
            mid_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_small:
            cur_idx += 1
            if cur_idx > max_count:
                break
            small_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_multi:
            cur_idx += 1
            if cur_idx > max_count:
                break
            multi_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_value:
            cur_idx += 1
            if cur_idx > max_count:
                break
            value_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_divi:
            cur_idx += 1
            if cur_idx > max_count:
                break
            divi_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        context["all_list"] = list(
            zip_longest(flex_list, large_list, lami_list, mid_list, small_list,
                        multi_list, value_list, divi_list, fillvalue='-'))

        return context

    def get_template_names(self):
        app_label = 'gmutfun'
        template_name_first = app_label + '/' + 'gmutfun_capbox_active_list.html'
        template_names_list = [template_name_first]
        return template_names_list


class GmutfunListView_Passive_Select_FOF_Capbox(ListView):
    model = Gmutfun

    # too many variants of 'NIFTY 50'
    # excluded hybrid funds by using , in it

    # queryset = q_gold | q_nifty | q_next | q_mid

    def get_queryset(self):

        self.queryset = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
            filter(Q(gmutfun_benchmark__contains='Domestic Price of Gold') | \
                   Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index') | \
                   Q(gmutfun_benchmark__contains='Next 50') |
                   Q(gmutfun_benchmark__contains='Midcap 150')). \
            exclude(gmutfun_benchmark__contains=',').order_by('gmutfun_benchmark', '-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        return self.queryset

    # @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(GmutfunListView_Passive_Select_FOF_Capbox, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url

        qs_nifty = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
            filter(Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_next = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
            filter(Q(gmutfun_benchmark__contains='Next 50')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_mid = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
            filter(Q(gmutfun_benchmark__contains='Midcap 150')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_gold = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
            filter(Q(gmutfun_benchmark__contains='Domestic Price of Gold')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_global = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
            exclude(gmutfun_benchmark__contains='Gold '). \
            exclude(gmutfun_benchmark__contains='NIFTY'). \
            exclude(gmutfun_benchmark__contains='Nifty'). \
            exclude(gmutfun_benchmark__contains='S&P BSE'). \
            exclude(gmutfun_benchmark__contains='Hybrid'). \
            exclude(gmutfun_benchmark__contains='Gilt'). \
            exclude(gmutfun_benchmark__contains='of Gold'). \
            exclude(gmutfun_benchmark__contains='Bond'). \
            exclude(gmutfun_benchmark__contains='REIT'). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_reit = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
            filter(Q(gmutfun_benchmark__contains='REIT')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        nifty_list = []
        next_list = []
        mid_list = []
        gold_list = []
        global_list = []
        reit_list = []

        # select at max 10
        max_count = 10

        cur_idx = 0
        for q1 in qs_nifty:
            cur_idx += 1
            if cur_idx > max_count:
                break
            nifty_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_next:
            cur_idx += 1
            if cur_idx > max_count:
                break
            next_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_mid:
            cur_idx += 1
            if cur_idx > max_count:
                break
            mid_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_gold:
            cur_idx += 1
            if cur_idx > max_count:
                break
            gold_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_global:
            cur_idx += 1
            if cur_idx > max_count:
                break
            global_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        cur_idx = 0
        for q1 in qs_reit:
            cur_idx += 1
            if cur_idx > max_count:
                break
            reit_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        context["all_list"] = list(
            zip_longest(nifty_list, next_list, mid_list, gold_list, global_list, reit_list, fillvalue='-'))

        return context

    def get_template_names(self):
        app_label = 'gmutfun'
        template_name_first = app_label + '/' + 'gmutfun_capbox_passive_list.html'
        template_names_list = [template_name_first]
        return template_names_list


class GmutfunListView_Passive_Select_ETF_Capbox(ListView):
    model = Gmutfun

    # too many variants of 'NIFTY 50'
    # excluded hybrid funds by using , in it

    # queryset = q_gold | q_nifty | q_next | q_mid

    def get_queryset(self):

        self.queryset = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='ETF')). \
            filter(Q(gmutfun_benchmark__contains='Domestic Price of Gold') | \
                   Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index') | \
                   Q(gmutfun_benchmark__contains='Next 50') |
                   Q(gmutfun_benchmark__contains='Midcap 150')). \
            exclude(gmutfun_benchmark__contains=',').order_by('gmutfun_benchmark', '-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        return self.queryset

    # @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(GmutfunListView_Passive_Select_ETF_Capbox, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url

        qs_nifty = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='ETF')). \
            filter(Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_next = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='ETF')). \
            filter(Q(gmutfun_benchmark__contains='Next 50')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_mid = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='ETF')). \
            filter(Q(gmutfun_benchmark__contains='Midcap 150')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_gold = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='ETF')). \
            filter(Q(gmutfun_benchmark__contains='Domestic Price of Gold')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_global = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='ETF')). \
            exclude(gmutfun_benchmark__contains='Gold'). \
            exclude(gmutfun_benchmark__contains='NIFTY'). \
            exclude(gmutfun_benchmark__contains='Nifty'). \
            exclude(gmutfun_benchmark__contains='S&P BSE'). \
            exclude(gmutfun_benchmark__contains='REIT'). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        qs_reit = Gmutfun.objects.all(). \
            filter(Q(gmutfun_type='ETF')). \
            filter(Q(gmutfun_benchmark__contains='REIT')). \
            exclude(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum'). \
            filter(gmutfun_aum__gte=100)

        nifty_list = []
        next_list = []
        mid_list = []
        gold_list = []
        global_list = []
        reit_list = []

        for q1 in qs_nifty:
            nifty_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        for q1 in qs_next:
            next_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        for q1 in qs_mid:
            mid_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        for q1 in qs_gold:
            gold_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        for q1 in qs_global:
            global_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        for q1 in qs_reit:
            reit_list.append(q1.gmutfun_scheme + ' - ' + str(int(q1.gmutfun_aum)))

        all_list = list(zip_longest(nifty_list, next_list, mid_list, gold_list, global_list, reit_list, fillvalue='-'))

        context["all_list"] = all_list

        return context

    def get_template_names(self):
        app_label = 'gmutfun'
        template_name_first = app_label + '/' + 'gmutfun_capbox_passive_list.html'
        template_names_list = [template_name_first]
        return template_names_list


class GmutfunListView_Passive_Select_ETF(ListView):
    model = Gmutfun

    # too many variants of 'NIFTY 50'
    # excluded hybrid funds by using , in it
    queryset = Gmutfun.objects.all(). \
        filter(Q(gmutfun_type='ETF')). \
        filter(Q(gmutfun_benchmark__contains='Domestic Price of Gold') | \
               Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index') | \
               Q(gmutfun_benchmark__contains='Next 50') |
               Q(gmutfun_benchmark__contains='Midcap 150')). \
        exclude(gmutfun_benchmark__contains=',').order_by('gmutfun_benchmark', '-gmutfun_aum'). \
        filter(gmutfun_aum__gte=100)

    # queryset = q_gold | q_nifty | q_next | q_mid

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Passive_Select_FOF(ListView):
    model = Gmutfun

    # too many variants of 'NIFTY 50'
    # excluded hybrid funds by using , in it
    queryset = Gmutfun.objects.all(). \
        filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        filter(Q(gmutfun_benchmark__contains='Domestic Price of Gold') | \
               Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index') | \
               Q(gmutfun_benchmark__contains='Next 50') |
               Q(gmutfun_benchmark__contains='Midcap 150')). \
        exclude(gmutfun_benchmark__contains=',').order_by('gmutfun_benchmark', '-gmutfun_aum'). \
        filter(gmutfun_aum__gte=100)

    # queryset = q_gold | q_nifty | q_next | q_mid

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context

class GmutfunListView_Nifty_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='Index').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Gold_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF').filter(gmutfun_benchmark__contains='Gold').order_by(
        '-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Gold_FOF(ListView):
    model = Gmutfun
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        filter(gmutfun_benchmark__contains='Gold').order_by(
        '-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Nifty_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF'). \
        filter(Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index')
               | Q(gmutfun_benchmark__contains='S&P BSE Sensex Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Next_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF'). \
        filter(gmutfun_benchmark__contains='Next 50').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Mid_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF'). \
        filter(gmutfun_benchmark__contains='Midcap 150').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Small_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF'). \
        filter(gmutfun_benchmark__contains='Smallcap 250').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context

class GmutfunListView_Global_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF'). \
        exclude(gmutfun_benchmark__contains='Gold'). \
        exclude(gmutfun_benchmark__contains='NIFTY'). \
        exclude(gmutfun_benchmark__contains='Nifty'). \
        exclude(gmutfun_benchmark__contains='S&P BSE').order_by('-gmutfun_aum')

    # filter(Q(gmutfun_benchmark__contains='MSCI')|Q(gmutfun_benchmark__contains='Russell')
    #       |Q(gmutfun_benchmark__contains='NASDAQ') | Q(gmutfun_benchmark__contains='Nasdaq')
    #       |Q(gmutfun_benchmark__contains='S&P 500') | Q(gmutfun_benchmark__contains='NYSE')
    #       |Q(gmutfun_benchmark__contains='Hang Seng')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Hybrid_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF'). \
        filter(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_NonGold_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(gmutfun_type='ETF'). \
        exclude(gmutfun_benchmark__contains='Gold').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Nifty_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        filter(Q(gmutfun_benchmark__contains='NIFTY 50 Total Return Index')
               | Q(gmutfun_benchmark__contains='S&P BSE Sensex Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Next_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        filter(gmutfun_benchmark__contains='Next 50').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Mid_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        filter(gmutfun_benchmark__contains='Midcap 150').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Small_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        filter(gmutfun_benchmark__contains='Smallcap 250').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context

class GmutfunListView_Global_FOF(ListView):
    model = Gmutfun

    # Used Gold with space to include Golden Dragon Index
    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        exclude(gmutfun_benchmark__contains='Gold '). \
        exclude(gmutfun_benchmark__contains='NIFTY'). \
        exclude(gmutfun_benchmark__contains='Nifty'). \
        exclude(gmutfun_benchmark__contains='S&P BSE'). \
        exclude(gmutfun_benchmark__contains='Hybrid'). \
        exclude(gmutfun_benchmark__contains='Gilt'). \
        exclude(gmutfun_benchmark__contains='of Gold'). \
        exclude(gmutfun_benchmark__contains='Bond'). \
        order_by('-gmutfun_aum')

    # filter(Q(gmutfun_benchmark__contains='MSCI') | Q(gmutfun_benchmark__contains='Russell')
    #       | Q(gmutfun_benchmark__contains='NASDAQ') | Q(gmutfun_benchmark__contains='Nasdaq')
    #       | Q(gmutfun_benchmark__contains='S&P 500')| Q(gmutfun_benchmark__contains='NYSE'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Hybrid_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all(). \
        filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        filter(gmutfun_benchmark__contains=',').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_NonGold_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all(). \
        filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        exclude(gmutfun_benchmark__contains='Gold').order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_Select(ListView):
    model = Gmutfun

    # too many variants of 'NIFTY 50'
    # excluded hybrid funds by using , in it

    # queryset = qs_flexi | qs_large | qs_mid | qs_small

    # queryset = q_gold | q_nifty | q_next | q_mid

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url

        score_grade = self.kwargs['score_grade']

        qs_flexi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_scheme__contains='Flexi')). \
            filter(Q(gmutfun_benchmark__contains='500 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_large = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='100 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_mid = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Midcap 150 Total Return Index') |
                   Q(gmutfun_benchmark__contains='150 MidCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        qs_small = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Smallcap 250 Total Return Index') |
                   Q(gmutfun_benchmark__contains='250 SmallCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_alpha_pct', '-gmutfun_aum')

        context["mf_flexi_list"] = qs_flexi
        context["mf_large_list"] = qs_large
        context["mf_mid_list"] = qs_mid
        context["mf_small_list"] = qs_small

        # using a fixed for all captype instead of different values by captype
        mf_captype_count = self.kwargs['mf_captype_count']
        # self.kwargs['flexi']
        # self.kwargs['large']
        # self.kwargs['mid']
        # self.kwargs['small']
        context["mf_flexi_count"] = mf_captype_count
        context["mf_large_count"] = mf_captype_count
        context["mf_mid_count"] = mf_captype_count
        context["mf_small_count"] = mf_captype_count

        return context

    def get_template_names(self):
        app_label = 'gmutfun'
        template_name_first = app_label + '/' + 'gmutfun_multi_list.html'
        template_names_list = [template_name_first]
        return template_names_list


class GmutfunListView_Active_AUM(ListView):
    model = Gmutfun

    # too many variants of 'NIFTY 50'
    # excluded hybrid funds by using , in it

    # queryset = qs_flexi | qs_large | qs_mid | qs_small

    # queryset = q_gold | q_nifty | q_next | q_mid

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url

        # cover all - grades
        score_grade = 0

        qs_flexi = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_scheme__contains='Flexi')). \
            filter(Q(gmutfun_benchmark__contains='500 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_aum')

        qs_large = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='100 Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_aum')

        qs_mid = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Midcap 150 Total Return Index') |
                   Q(gmutfun_benchmark__contains='150 MidCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_aum')

        qs_small = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
            filter(Q(gmutfun_benchmark__contains='Smallcap 250 Total Return Index') |
                   Q(gmutfun_benchmark__contains='250 SmallCap Total Return Index')). \
            exclude(gmutfun_benchmark__contains=','). \
            filter(gmutfun_aum__gte=1000). \
            filter(gmutfun_alpha_grade__gte=score_grade). \
            order_by('-gmutfun_aum')

        context["mf_flexi_list"] = qs_flexi
        context["mf_large_list"] = qs_large
        context["mf_mid_list"] = qs_mid
        context["mf_small_list"] = qs_small

        # using a fixed for all captype instead of different values by captype
        mf_captype_count = self.kwargs['mf_captype_count']
        # self.kwargs['flexi']
        # self.kwargs['large']
        # self.kwargs['mid']
        # self.kwargs['small']
        context["mf_flexi_count"] = mf_captype_count
        context["mf_large_count"] = mf_captype_count
        context["mf_mid_count"] = mf_captype_count
        context["mf_small_count"] = mf_captype_count

        return context

    def get_template_names(self):
        app_label = 'gmutfun'
        template_name_first = app_label + '/' + 'gmutfun_multi_list.html'
        template_names_list = [template_name_first]
        return template_names_list


class GmutfunListView_Active_Flexi(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_scheme__contains='Flexi')). \
        filter(Q(gmutfun_benchmark__contains='500 Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_Large(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_benchmark__contains='100 Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_Mid(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_benchmark__contains='Midcap 150 Total Return Index') |
               Q(gmutfun_benchmark__contains='150 MidCap Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_Small(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_benchmark__contains='Smallcap 250 Total Return Index') |
               Q(gmutfun_benchmark__contains='250 SmallCap Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_Multi(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_benchmark__contains='Nifty 500 Multicap')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_LargeMid(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_benchmark__contains='Large Midcap 250 Total Return Index') |
               Q(gmutfun_benchmark__contains='250 Large MidCap Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_Value(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_scheme__contains='Value')). \
        filter(Q(gmutfun_benchmark__contains='500 Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Active_Dividend(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Active')). \
        filter(Q(gmutfun_scheme__contains='Dividend')). \
        filter(Q(gmutfun_benchmark__contains='500 Total Return Index')).order_by('-gmutfun_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Gmutfun_url()
        context["refresh_url"] = refresh_url
        return context


class GmutfunListView_Benchmark_FOF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='Index') | Q(gmutfun_type='FoF')). \
        values('gmutfun_benchmark').annotate(scheme_count=Count('gmutfun_benchmark')). \
        order_by('-scheme_count')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # get count of Industries
        benchmarks_count = len(Gmutfun.objects.all().values('gmutfun_benchmark'). \
                               annotate(benchmarks_count=Count('gmutfun_benchmark', distinct=True)))
        context['benchmarks_count'] = benchmarks_count
        return context


class GmutfunListView_Benchmark_ETF(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().filter(Q(gmutfun_type='ETF')).values('gmutfun_benchmark').annotate(
        scheme_count=Count('gmutfun_benchmark')). \
        order_by('-scheme_count')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # get count of Industries
        benchmarks_count = len(Gmutfun.objects.all().values('gmutfun_benchmark'). \
                               annotate(benchmarks_count=Count('gmutfun_benchmark', distinct=True)))
        context['benchmarks_count'] = benchmarks_count
        return context


class GmutfunListView_Benchmark_All(ListView):
    model = Gmutfun

    queryset = Gmutfun.objects.all().values('gmutfun_benchmark').annotate(scheme_count=Count('gmutfun_benchmark')). \
        order_by('-scheme_count')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # get count of Industries
        benchmarks_count = len(Gmutfun.objects.all().values('gmutfun_benchmark'). \
                               annotate(benchmarks_count=Count('gmutfun_benchmark', distinct=True)))
        context['benchmarks_count'] = benchmarks_count
        return context

def Gmutfun_url():
    url = 'https://archives.nseindia.com/content/umufub/ind_nifty500list.csv'

    return url


# one parameter named request
def Gmutfun_fetch(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()
    debug_level = 1
    print('fetch not supported')
    return HttpResponseRedirect(reverse("gmutfun-list"))


# one parameter named request
def Gmutfun_upload(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    template = "gmutfun/gmutfun_list.html"
    data = Gmutfun.objects.all()

    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template)

    # delete existing records
    print('Deleted existing Gmutfun data')
    Gmutfun.objects.all().delete()

    for req_file in request.FILES.getlist("files"):

        if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
            # get worksheet name
            # print('temporary file path:', req_file.temporary_file_path)
            print(req_file)

            if True:
                wb = openpyxl.load_workbook(req_file)
                print(wb.sheetnames)
                sheet_name = wb.sheetnames[0]
                print(sheet_name)
                ws = wb[sheet_name]
                df = pd.DataFrame(ws.values)
            else:
                xl = pd.ExcelFile(req_file)
                if debug_level > 0:
                    print(xl.sheet_names)
                # single worksheet - Data
                sheet_name = xl.sheet_names[0]
                df = xl.parse(sheet_name)

            # can be 'Data'
            # can be 'Average MCap Jan Jun 2020'
            # if sheet_name != 'fund-performance':
            #    print("sheet name changed to", sheet_name)

            # ignore top 6 line : Value Research, Fund Performance
            # remove top six line from dataframe
            df = df.iloc[6:]

            if debug_level > 0:
                print("old columns : ")
                print(df.columns)

            # change column name of data frame
            columns_list = ["scheme_name", "benchmark", "risk_scheme", "risk_benchmark", "nav_date", "nav_regular",
                            "nav_direct",
                            "return_1y_pct_regular", "return_1y_pct_direct", "return_1y_pct_benchmark",
                            "return_3y_pct_regular", "return_3y_pct_direct", "return_3y_pct_benchmark",
                            "return_5y_pct_regular", "return_5y_pct_direct", "return_5y_pct_benchmark",
                            "return_10y_pct_regular", "return_10y_pct_direct", "return_10y_pct_benchmark",
                            "return_since_launch_regular", "return_since_launch_direct",
                            "return_since_launch_benchmark",
                            "daily_aum"]

            df.columns = columns_list

            if debug_level > 0:
                print("new columns : ")
                print(df.columns)

            # Keep only top 1000 entries
            # df = df.iloc[:1000]

            # drop columns that are not required
            skip_columns_list = ["risk_scheme", "risk_benchmark", "nav_date", "nav_regular", "nav_direct",
                                 "return_1y_pct_direct", "return_3y_pct_direct",
                                 "return_5y_pct_direct", "return_10y_pct_direct",
                                 "return_since_launch_regular", "return_since_launch_direct",
                                 "return_since_launch_benchmark", ]

            # removed from skip list
            #                                 "return_1y_pct_regular", "return_1y_pct_direct", "return_1y_pct_benchmark",
            #                                 "return_3y_pct_regular", "return_3y_pct_direct", "return_3y_pct_benchmark",
            #                                 "return_5y_pct_regular", "return_5y_pct_direct", "return_5y_pct_benchmark",
            #                                 "return_10y_pct_regular", "return_10y_pct_direct", "return_10y_pct_benchmark",

            df.drop(skip_columns_list, axis=1, inplace=True)

            if debug_level > 0:
                print(df)

            # replace empty value with NaN
            df['daily_aum'].replace('', np.nan, inplace=True)

            # drop empty values
            df.dropna(subset=['daily_aum'], inplace=True)

            # round avg_mcap
            # df = df.round({'avg_mcap' : 1})
            # covert to numeric
            # df[["avg_mcap"]] = df[["avg_mcap"]].apply(pd.to_numeric)
            df[["daily_aum"]] = df[["daily_aum"]].astype(int)

            data_set = df.to_csv(header=True, index=False)

        if req_file.name.endswith('.csv'):
            data_set = req_file.read().decode('UTF-8')

        if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx')):
            messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV FILE.')
            return HttpResponseRedirect(reverse("gmutfun-list"))

        # setup a stream which is when we loop through each line we are able to handle a data in a stream

        io_string = io.StringIO(data_set)
        # skip top three rows
        next(io_string)

        skip_records = 0
        for column in csv.reader(io_string, delimiter=',', quotechar='"'):

            gmutfun_scheme = column[0].strip()
            gmutfun_benchmark = column[1].strip()

            if re.search('Index', gmutfun_scheme):
                gmutfun_type = 'Index'
            elif re.search('ETF', gmutfun_scheme):
                gmutfun_type = 'ETF'
            elif re.search('Flexi', gmutfun_scheme):
                gmutfun_type = 'Active'
            elif re.search('Large', gmutfun_scheme) \
                    or re.search('Bluechip', gmutfun_scheme) \
                    or re.search('Top 100', gmutfun_scheme):
                gmutfun_type = 'Active'
            elif re.search('Mid', gmutfun_scheme):
                gmutfun_type = 'Active'
            elif re.search('Small', gmutfun_scheme) or re.search('Emerging', gmutfun_scheme):
                gmutfun_type = 'Active'
            elif re.search('Multi', gmutfun_scheme):
                gmutfun_type = 'Active'
            elif re.search('Value', gmutfun_scheme):
                gmutfun_type = 'Active'
            elif re.search('Dividend', gmutfun_scheme):
                gmutfun_type = 'Active'
            elif re.search('FoF', gmutfun_scheme, re.IGNORECASE):
                gmutfun_type = 'FoF'
            else:
                if re.search('Large Midcap', gmutfun_benchmark):
                    gmutfun_type = 'Active'
                elif re.search('Large', gmutfun_benchmark):
                    gmutfun_type = 'Active'
                elif re.search('Midcap', gmutfun_benchmark):
                    gmutfun_type = 'Active'
                elif re.search('Multi', gmutfun_benchmark):
                    gmutfun_type = 'Active'
                else:
                    gmutfun_type = 'Unk'

            gmutfun_subtype = ''
            if gmutfun_type == 'Active':
                gmutfun_subtype = comfun.comm_mf_subcat(gmutfun_scheme, gmutfun_benchmark)

            gmutfun_ret_1y_reg = column[2].strip()
            gmutfun_ret_1y_bench = column[3].strip()

            gmutfun_ret_3y_reg = column[4].strip()
            gmutfun_ret_3y_bench = column[5].strip()

            gmutfun_ret_5y_reg = column[6].strip()
            gmutfun_ret_5y_bench = column[7].strip()

            gmutfun_ret_10y_reg = column[8].strip()
            gmutfun_ret_10y_bench = column[9].strip()

            if gmutfun_ret_1y_reg == '':
                gmutfun_ret_1y_reg = '0'
            if gmutfun_ret_1y_bench == '':
                gmutfun_ret_1y_bench = '0'

            if gmutfun_ret_3y_reg == '':
                gmutfun_ret_3y_reg = '0'
            if gmutfun_ret_3y_bench == '':
                gmutfun_ret_3y_bench = '0'

            if gmutfun_ret_5y_reg == '':
                gmutfun_ret_5y_reg = '0'
            if gmutfun_ret_5y_bench == '':
                gmutfun_ret_5y_bench = '0'

            if gmutfun_ret_10y_reg == '':
                gmutfun_ret_10y_reg = '0'
            if gmutfun_ret_10y_bench == '':
                gmutfun_ret_10y_bench = '0'

            gmutfun_aum = column[10].strip()

            gmutfun_ret_1y_reg = round(float(gmutfun_ret_1y_reg), 2)
            gmutfun_ret_3y_reg = round(float(gmutfun_ret_3y_reg), 2)
            gmutfun_ret_5y_reg = round(float(gmutfun_ret_5y_reg), 2)
            gmutfun_ret_10y_reg = round(float(gmutfun_ret_10y_reg), 2)

            gmutfun_ret_1y_bench = round(float(gmutfun_ret_1y_bench), 2)
            gmutfun_ret_3y_bench = round(float(gmutfun_ret_3y_bench), 2)
            gmutfun_ret_5y_bench = round(float(gmutfun_ret_5y_bench), 2)
            gmutfun_ret_10y_bench = round(float(gmutfun_ret_10y_bench), 2)

            gmutfun_alpha_grade = 0

            # exclude 0.0 data
            if int(gmutfun_ret_1y_reg) != 0 and math.ceil(gmutfun_ret_1y_reg) >= math.floor(gmutfun_ret_1y_bench):
                gmutfun_alpha_grade += 1
            if int(gmutfun_ret_3y_reg) != 0 and math.ceil(gmutfun_ret_3y_reg) >= math.floor(gmutfun_ret_3y_bench):
                gmutfun_alpha_grade += 1
                print('score ', gmutfun_scheme, gmutfun_ret_3y_reg, gmutfun_ret_3y_bench)
            if int(gmutfun_ret_5y_reg) != 0 and math.ceil(gmutfun_ret_5y_reg) >= math.floor(gmutfun_ret_5y_bench):
                gmutfun_alpha_grade += 1
                print('score ', gmutfun_scheme, gmutfun_ret_5y_reg, gmutfun_ret_5y_bench)
            if int(gmutfun_ret_10y_reg) != 0 and math.ceil(gmutfun_ret_10y_reg) >= math.floor(gmutfun_ret_10y_bench):
                gmutfun_alpha_grade += 1
                print('score ', gmutfun_scheme, gmutfun_ret_10y_reg, gmutfun_ret_10y_bench)

            gmutfun_alpha_pct = 0
            gmutfun_alpha_pct += (gmutfun_ret_1y_reg - gmutfun_ret_1y_bench)
            gmutfun_alpha_pct += (gmutfun_ret_3y_reg - gmutfun_ret_3y_bench)
            gmutfun_alpha_pct += (gmutfun_ret_5y_reg - gmutfun_ret_5y_bench)
            gmutfun_alpha_pct += (gmutfun_ret_10y_reg - gmutfun_ret_10y_bench)

            _, created = Gmutfun.objects.update_or_create(
                gmutfun_scheme=gmutfun_scheme,
                gmutfun_type=gmutfun_type,
                gmutfun_subtype=gmutfun_subtype,
                gmutfun_benchmark=gmutfun_benchmark,
                gmutfun_ret_1y_reg=gmutfun_ret_1y_reg,
                gmutfun_ret_1y_bench=gmutfun_ret_1y_bench,
                gmutfun_ret_3y_reg=gmutfun_ret_3y_reg,
                gmutfun_ret_3y_bench=gmutfun_ret_3y_bench,
                gmutfun_ret_5y_reg=gmutfun_ret_5y_reg,
                gmutfun_ret_5y_bench=gmutfun_ret_5y_bench,
                gmutfun_ret_10y_reg=gmutfun_ret_10y_reg,
                gmutfun_ret_10y_bench=gmutfun_ret_10y_bench,
                gmutfun_aum=gmutfun_aum,
                gmutfun_alpha_grade=gmutfun_alpha_grade,
                gmutfun_alpha_pct=gmutfun_alpha_pct
            )

    lastrefd_update("gmutfun")

    print('Skipped records', skip_records)
    print('Completed loading new Gmutfun data')
    return HttpResponseRedirect(reverse("gmutfun-list"))
